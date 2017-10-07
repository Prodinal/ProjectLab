import sys
import evaluate_image_slice
import tensorflow as tf
import cv2
import heapq
import datetime
import openpyxl
import os

#classes
class resultObj(object):
    accuracy = -1
    sensitivity = -1
    specificity = -1
    true_pos = []
    false_pos = []
    true_neg = []
    false_neg = []
    undecided = []
    parameters = ""

    def __init__(self):
        self.date = datetime.datetime.now()

    def calculateIndicators(self):
        true_pos = len(self.true_pos)
        true_neg = len(self.true_neg)
        false_pos = len(self.false_pos)
        false_neg = len(self.false_neg)
        self.accuracy = (true_pos + true_neg) / (true_neg + true_pos + false_neg + false_pos)
        if true_pos + false_neg != 0:
            self.sensitivity = true_pos / (true_pos + false_neg)
        if true_neg + false_pos != 0:
            self.specificity = true_neg / (true_neg + false_pos)
    
    def saveToFile(self):
        # append data to file, create if does not exist, add date too
        wb = openpyxl.load_workbook(resultDocumentsPath + "Results.xlsx")
        ws = wb.active
        col = ws.max_column + 1
        ws.cell(row=1, column=col).value = self.date
        ws.cell(row=2, column=col).value = self.accuracy
        ws.cell(row=3, column=col).value = self.sensitivity
        ws.cell(row=4, column=col).value = self.specificity
        ws.cell(row=5, column=col).value = self.parameters
        ws.cell(row=6, column=col).value = len(self.true_pos)
        ws.cell(row=7, column=col).value = len(self.false_pos)
        ws.cell(row=8, column=col).value = len(self.true_neg)
        ws.cell(row=9, column=col).value = len(self.false_neg)
        ws.cell(row=10, column=col).value = len(self.undecided)
                
        wb.save(resultDocumentsPath + "Results.xlsx")

#constants
winW = 298
winH = 298
# how close the tumor has to be to the middle of the boundingbox in order for it to be positive ( % of the window size)
# 0.4 is 40% of the square cut away from each side, e.g. the middle 0.2WinW by 0.2WinH square
errorTolerance = 0.4
resultDocumentsPath = 'd:\\School\\2017Onlab\\ProjectLab\\Code\\GoogleNet\\Results\\'
#
clone = None


#functions
def sliding_window(image, stepSize, windowSize):
    # slide a window across the image
    for y in range(0, image.shape[0], stepSize):
        for x in range(0, image.shape[1], stepSize):
            # yield the current window
            yield (x, y, image[y:y + windowSize[1], x:x + windowSize[0]])


def getCoordsFromLine(line):
        parts = line.split('	')
        return int(parts[5]), int(parts[6])


def getFileNameFromLine(line):
    parts = line.split('	')
    return parts[0]

def getCorrectCoords(file, name):
    coordsOfCorrect = None
    for line in file:
        if getFileNameFromLine(line).split('.')[0] == name.split('.')[0]:
            return getCoordsFromLine(line)


def getFileNameFromImagePath(path):
    return os.path.basename(path)


def calculateAccuracy(expectedCoord, tumors, healthy):
    # calculate accuracy, specificity, and so on
    epsilon_x = winW * errorTolerance
    epsilon_y = winH * errorTolerance
    
    result = resultObj()
    true_pos = 0
    true_neg = 0
    false_pos = 0
    false_neg = 0

    x,y = expectedCoord
    for ((bot_x, bot_y),(top_x,top_y)), score in tumors:
        # the tumor is inside the rectangle
        if x >= bot_x + epsilon_x and x <= top_x - epsilon_x and \
           y >= bot_y + epsilon_y and y <= top_y - epsilon_y:
            true_pos += 1
        else:
            false_pos += 1
    for ((bot_x, bot_y),(top_x,top_y)), score in healthy:
        # the tumor is inside the rectangle
        if x >= bot_x + epsilon_x and x <= top_x - epsilon_x and \
           y >= bot_y + epsilon_y and y <= top_y - epsilon_y:
            false_neg += 1
        else:
            true_neg += 1
    
    print("true positive: " + str(true_pos))
    print("false positive: " + str(false_pos))
    print("true negative: " + str(true_neg))
    print("false negative: " + str(false_neg))
    result.accuracy = (true_pos + true_neg) / (true_neg + true_pos + false_neg + false_pos)
    if true_pos + false_neg != 0:
        result.sensitivity = true_pos / (true_pos + false_neg)
    if true_neg + false_pos != 0:
        result.specificity = true_neg / (true_neg + false_pos)

    print('Accuracy: ' + str(result.accuracy))
    print('Sensitivity: ' + str(result.sensitivity))
    print('Specificity: ' + str(result.specificity))

    return result
    


def saveAccuracyToFile(data):
    # append data to file, create if does not exist, add date too
    wb = openpyxl.load_workbook(resultDocumentsPath + "Results.xlsx")
    ws = wb.active
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = data.date
    ws.cell(row=2, column=col).value = data.accuracy
    ws.cell(row=3, column=col).value = data.sensitivity
    ws.cell(row=4, column=col).value = data.specificity
    ws.cell(row=5, column=col).value = "parameters"
    
    wb.save(resultDocumentsPath + "Results.xlsx")

def getPixelOfLungMaskFromImageCoord(x,y, mask):
    maskClone = mask.copy()
    cv2.circle(maskClone, (int(x/2), int(y/2)), 4, (255, 127, 0))
    cv2.imshow("LungMask", maskClone)
    cv2.waitKey(1)
    pixel = mask[int(y/2), int(x/2)]
    return pixel[0]

#Main
def Main():
    if len(sys.argv) < 2:
        print("Too few arguments, reverting to default image")
        # sys.exit(-1)

    label_lines = [line.rstrip() for line
                in tf.gfile.GFile("d:/tmp/output_labels.txt")]

    evaluate_image_slice.init_tf("D:\\tmp\\output_graph.pb", label_lines)

    folder_path = ("d:\\School\\2017Onlab\\ProjectLab\\Code\\"
                "Creating_Training_Images\\EveryImage\\")
    image_name = "JPCLN001Result.png"
    image_path = folder_path + image_name

    if len(sys.argv) >= 2:
        image_path = sys.argv[1]

    # image_data = tf.gfile.FastGFile(image_path, 'rb').read()
    image_data = cv2.imread(image_path)

    info_file = open(os.path.join(
                    folder_path,
                    'CLNDAT_E.txt'),
                'r')
    coordsOfCorrect = getCorrectCoords(info_file, getFileNameFromImagePath(image_path).replace('Result', ''))    

    lungmask_path = "d:\\School\\2017Onlab\\ProjectLab\\Code\\Creating_Training_Images\\EveryImage\\LungMasks\\" + getFileNameFromImagePath(image_path).replace('Result', '')
    lungmask_data = cv2.imread(lungmask_path)
    # lungmask_data = cv2.cvtColor(lungmask_data, cv2.COLOR_BGR2GRAY)

    tumors = []
    healthy = []
    cantTell = []

    idx = 0
    for (x, y, window) in sliding_window(image_data, 32, (winW, winH)):
            try:
                # if the window does not meet our desired window size, ignore it
                if window.shape[0] != winH or window.shape[1] != winW:
                    continue
                
                # if the middle of the sliding window is not inside the lung area (based on lung mask) ignore it
                pixelOfLung = getPixelOfLungMaskFromImageCoord(x + int(winW/2), y + int(winH/2), lungmask_data)
                if pixelOfLung == 0:
                    continue

                predictions = evaluate_image_slice.evaluateSlice(window)
                print(predictions)

                # positive > negative
                if predictions[0][0] > 0.9: # 0.6
                    # tumor format is ( ((bottom left x,y), (top right x,y)), score)
                    tumors.append((((x, y), (x+winW, y+winH)), predictions[0][0]))
                elif predictions[0][0] < 0.4:
                    healthy.append((((x, y), (x+winW, y+winH)), predictions[0][0]))
                else:
                    cantTell.append((((x, y), (x+winW, y+winH)), predictions[0][0]))

                if (idx % 10 == 0) or True:
                    clone = image_data.copy()

                    top_10_tumors = heapq.nlargest(10, tumors, key=lambda t: t[1])

                    for ((tumor_xbl, tumor_ybl), (tumor_xtr, tumor_ytr)), _ in tumors:
                        cv2.rectangle(clone, (tumor_xbl, tumor_ybl), (tumor_xtr, tumor_ytr), (0, 0, 255), 2)

                    cv2.rectangle(clone, (x, y), (x + winW, y + winH), color=(0, 255, 0), thickness=2)
                    clone = cv2.resize(clone, (0,0), fx=0.5, fy=0.5)
                    cv2.imshow("Window", clone)
                    cv2.waitKey(1)
                
                idx += 1
            except KeyboardInterrupt:
                break
    

    accuracy_data = calculateAccuracy(coordsOfCorrect, tumors, healthy)
    saveAccuracyToFile(accuracy_data)
    
    if clone is not None:
        cv2.imwrite(resultDocumentsPath + image_path[-12:], clone)
    else:
        print("Clone of image is null :c")
    #save image to resultDocumentsPath
    print("End of program")
    # evaluate_image_slice.evaluateSlice(label_lines, image_data)

    # image_name = "JPCLN069Nodule298.jpg"
    # image_path = folder_path + image_name
    # image_data = tf.gfile.FastGFile(image_path, 'rb').read()

    # evaluate_image_slice.evaluateSlice(label_lines, image_data)
if __name__ == "__main__":
    Main()